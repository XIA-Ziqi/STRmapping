#!/usr/bin/python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook
import kdtree
import math
import re
import pickle
import urllib
from urllib import request, parse
from bs4 import BeautifulSoup
import http.cookiejar

class Item(object):
    def __init__(self, x, y, title, url, address):
        self.coords = (x, y)
        self.title = title
        self.url = url
        self.address = address

    def __len__(self):
        return len(self.coords)

    def __getitem__(self, i):
        return self.coords[i]

    def __repr__(self):
        return 'Item({}, {}, {},{},{})'.format(self.coords[0], self.coords[1], self.title, self.url, self.address)

x_pi = 3.14159265358979324 * 3000.0 / 180.0
pi = 3.1415926535897932384626  # π
a = 6378245.0  # 长半轴
ee = 0.00669342162296594323  # 偏心率平方

def baidusearch(url_key):
    #书写完整url
    url = "http://www.baidu.com/s?wd="
    key = url_key
    key_code = urllib.parse.quote(key)
    url_all = url + key_code

    #添加headers
    headers = {
    "User-Agent": "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.81 Safari/537.36",
    }

    #读取响应并打印
    request = urllib.request.Request(url_all,headers=headers)
    response = urllib.request.urlopen(request)
    data= response.read().decode('utf-8')
    soup = BeautifulSoup(data,'html.parser') #使用python标准库来作为解析器'html.parser'  也可以"lxml"，"html5lib"
    i=0
    url_list=[]
    for result_table in soup.find_all('div',class_="result c-container"): #, class_='t'可以不加
        
        if i<6:
            a_click = result_table.find("a")
            print("_________________________________________________________________________")
            print(str(i),"-----标题----" + a_click.get_text())  # 标题
            div_click =result_table.find('div',class_="c-abstract")
            print("简介:" + div_click.get_text())  # 标题
            if "携程" in key:
                url_list.append(str(a_click.get("href")))
                print("链接:" + str(a_click.get("href")))  # 链接
        i=i+1

    choose=int(input("num? "))
    if choose in range(0,6):
        return(urllib.request.urlopen(url_list[choose]).geturl())

def gcj02_to_bd09(lng, lat):
    """
    火星坐标系(GCJ-02)转百度坐标系(BD-09)
    谷歌、高德——>百度
    :param lng:火星坐标经度
    :param lat:火星坐标纬度
    :return:
    """
    z = math.sqrt(lng * lng + lat * lat) + 0.00002 * math.sin(lat * x_pi)
    theta = math.atan2(lat, lng) + 0.000003 * math.cos(lng * x_pi)
    bd_lng = z * math.cos(theta) + 0.0065
    bd_lat = z * math.sin(theta) + 0.006
    return [bd_lng, bd_lat]

# 读取数据
print("reading data...")
pickle_file = 'sql.pickle'
with open(pickle_file, 'rb') as f:
    pickle_data = pickle.load(f)       # 反序列化，与pickle.dump相反
    points = pickle_data['pints']
    del pickle_data  # 释放内存
# 把携程的数据建成树
print("start building tree...")
tree = kdtree.create(points)
print("Finished building 2d tree, load: %d items" % (len(points)))

# 读取str的数据
print("Prepare to load xlsx file")
wb = load_workbook(filename='8dc13e6f1ef9d136.xlsx')
ws = wb['independent']
print("Finished loading xlsx file")
ws1 = wb['result_independent ']
ws1.cell(row=3, column=1).value = 'str_id'
ws1.cell(row=3, column=2).value = 'str_name'
ws1.cell(row=3, column=3).value = 'str_address1'
ws1.cell(row=3, column=4).value = "ctrip_title"
ws1.cell(row=3, column=5).value = "ctrip_address"
ws1.cell(row=3, column=6).value = "ctrip_url"


def write_data(row, data):
    ws1.cell(row=row, column=4).value = data["title"]
    ws1.cell(row=row, column=5).value = data["address"]
    ws1.cell(row=row, column=6).value = data["url"]


def item_to_dict(item: Item):
    return {"title": item.title, "address": item.address, "url": item.url}


k = 35
count = 0
#############################修改此处数字################################################################################
for row in range(1300,1305):
######################################################################################################################
    obj = {}
    obj['id'] = ws.cell(row=row, column=2).value
    obj['url'] = ws.cell(row=row, column=3).value
    obj['name'] = ws.cell(row=row, column=4).value
    obj['address1'] = ws.cell(row=row, column=5).value
    obj['address2'] = ws.cell(row=row, column=7).value
    obj['address_ctrip'] = ws.cell(row=row, column=8).value
    obj['latitude'] = ws.cell(row=row, column=14).value
    obj['longitude'] = ws.cell(row=row, column=15).value
    ws1.cell(row=row, column=1).value = obj['id']
    ws1.cell(row=row, column=2).value = obj['name']
    ws1.cell(row=row, column=3).value = obj['address1']

    if (not isinstance(obj['latitude'], float) or not isinstance(obj['longitude'], float)):
        print("\nerror data, skip...")
        continue
    try:
        longitude, latitude = gcj02_to_bd09(obj['longitude'], obj['latitude'])
        resourses = tree.search_knn([longitude, latitude], k)  # 最近的点位（是一个list）
        print('第'+str(row)+'个参考地址：', obj['address1'], '参考名称:', obj['name'])
        str_num_address = "".join(re.findall(r"\d+", obj['address1']))
        ans = 0

        for i in range(1, k+1):
            if i < 36:
                print(
                    str(i), ':', resourses[i-1][0].data.address, '名称：', resourses[i-1][0].data.title)
            if str_num_address != "":
                addressi = resourses[i-1][0].data.address
                if addressi == None:
                    continue

                ctrip_num_address = re.findall(
                    r"\d+", resourses[i-1][0].data.address)

                for j in range(1, len(ctrip_num_address)+1):
                    if (str_num_address == ctrip_num_address[j-1] and len(str_num_address) != 0):
                        if ans == 0:
                            ans = i
                        print(
                            '-------------------------------------------------------')
                        count += 1
                        break

        print("Is", str(ans), "the right position y/c/n?")
        ansyn = input()
        if ans != 0 and ansyn == 'y':
            write_data(row, item_to_dict(resourses[ans-1][0].data))
        elif ansyn == 'c':
            print("input the number you want: ")
            selected = int(input())
            if selected > 0 and selected <= k:
                write_data(row, item_to_dict(resourses[selected-1][0].data))
            else:
                print("out of range in resourses")
        else:
            url=0
            print('第'+str(row)+'个参考地址：', obj['address1'], '参考名称:', obj['name'])
            url_key0='携程'+obj['name']
            #url_key=obj['name']
            url=baidusearch(url_key0)
            if url== None:
                print('第'+str(row)+'个参考地址：', obj['address1'], '参考名称:', obj['name'])
                url_key22 = input('input the right name in chinese:')
                url_key2='携程'+url_key22
                url=baidusearch(url_key2)
            print(url)
            if url==None:
                url = input('input the right url:')
            write_data(row, {"title": 'Null', "address": 'Null', "url": url})


        print("________________________________________________________________")

    except Exception as e:
        print("error in search nn: \n")
        print(e)

print('\n writing to output.xlsx')

wb.save(filename="8dc13e6f1ef9d136.xlsx")
print("finished!!!!!!")
print("automated find", str(count))
